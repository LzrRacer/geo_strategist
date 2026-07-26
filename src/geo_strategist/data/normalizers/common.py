"""Shared helpers for deterministic Excel normalization."""

from __future__ import annotations

import json
import warnings as warning_lib
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from geo_strategist.data.inventory import sha256_file
from geo_strategist.data.normalization import (
    ExtractionMapping,
    MappingStatus,
    NormalizationManifest,
    NormalizedRecord,
    SourceTable,
    infer_value_type,
    normalize_cell_value,
    normalize_field_name,
    safe_text,
    source_table_id,
)
from geo_strategist.data.provenance import ProvenanceRecord, SourceKind, SourceRef


@dataclass(frozen=True)
class HeaderDetection:
    """Conservative header-detection result."""

    status: MappingStatus
    header_row: int | None
    confidence: float
    column_names: list[str]
    value_columns: list[int]
    label_columns: list[int]
    year_columns: list[int]
    age_columns: list[int]
    warnings: list[str]


def load_yaml(path: str | Path) -> dict[str, Any]:
    """Load a YAML config file."""

    with Path(path).open("r", encoding="utf-8") as file:
        loaded = yaml.safe_load(file)
    return loaded or {}


def read_inspection_profile(path: Path) -> dict[str, Any] | None:
    """Load an inspection profile when present."""

    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: Any) -> None:
    """Write JSON with stable formatting."""

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_jsonl(path: Path, records: list[NormalizedRecord]) -> None:
    """Write normalized records as JSONL."""

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "\n".join(record.model_dump_json() for record in records)
        + ("\n" if records else ""),
        encoding="utf-8",
    )


def write_markdown_report(
    path: Path,
    title: str,
    source_tables: list[SourceTable],
    mappings: list[ExtractionMapping],
    record_count: int,
    warnings: list[str],
) -> None:
    """Write a concise mapping markdown report."""

    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        f"# {title}",
        "",
        f"- Source tables: {len(source_tables)}",
        f"- Normalized records: {record_count}",
        f"- Unresolved mappings: {sum(1 for m in mappings if m.status is MappingStatus.UNRESOLVED)}",
        "",
        "| Sheet | Status | Confidence | Header row | Value columns | Label columns |",
        "| --- | --- | ---: | ---: | ---: | ---: |",
    ]
    for mapping in mappings:
        lines.append(
            f"| {mapping.sheet_name} | {mapping.status.value} | {mapping.confidence:.2f} | "
            f"{mapping.header_row or ''} | {len(mapping.value_columns)} | "
            f"{len(mapping.label_columns)} |"
        )
    if warnings:
        lines.extend(["", "## Warnings"])
        lines.extend(f"- {warning}" for warning in warnings)
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def workbook_with_warnings(path: Path) -> tuple[Any, list[str]]:
    """Open a workbook while collecting nonfatal parser warnings."""

    with warning_lib.catch_warnings(record=True) as caught_warnings:
        warning_lib.simplefilter("always", UserWarning)
        workbook = load_workbook(path, read_only=True, data_only=False)
    warnings = sorted({str(warning.message) for warning in caught_warnings})
    return workbook, warnings


def row_values(sheet: Any, row_index: int) -> list[Any]:
    """Return values for one worksheet row."""

    return list(
        next(
            sheet.iter_rows(
                min_row=row_index,
                max_row=row_index,
                max_col=sheet.max_column,
                values_only=True,
            ),
            (),
        )
    )


def worksheet_values(sheet: Any) -> list[list[Any]]:
    """Read worksheet values once to avoid repeated XML scans in read-only mode."""

    return [list(row) for row in sheet.iter_rows(values_only=True)]


def values_row(rows: list[list[Any]], row_index: int, column_count: int) -> list[Any]:
    """Return a 1-indexed row from cached worksheet values."""

    if row_index < 1 or row_index > len(rows):
        return []
    row = rows[row_index - 1]
    if len(row) < column_count:
        return [*row, *([None] * (column_count - len(row)))]
    return row[:column_count]


def is_blank(value: Any) -> bool:
    """Return whether a source cell should be treated as blank."""

    return value is None or (isinstance(value, str) and not value.strip())


def detect_generic_header(
    sheet: Any,
    max_header_scan_rows: int,
    min_non_empty_header_cells: int,
    min_data_rows_after_header: int,
    min_confidence_for_extraction: float,
    rows: list[list[Any]] | None = None,
) -> HeaderDetection:
    """Detect a simple row-oriented table header without semantic guessing."""

    cached_rows = rows if rows is not None else worksheet_values(sheet)
    row_count = len(cached_rows)
    column_count = int(sheet.max_column or 0)
    if row_count == 0 or column_count == 0:
        return HeaderDetection(
            status=MappingStatus.UNRESOLVED,
            header_row=None,
            confidence=0,
            column_names=[],
            value_columns=[],
            label_columns=[],
            year_columns=[],
            age_columns=[],
            warnings=["Sheet has no usable dimensions."],
        )

    best: tuple[int, float, list[str], list[int]] | None = None
    for row_index in range(1, min(row_count, max_header_scan_rows) + 1):
        values = values_row(cached_rows, row_index, column_count)
        non_empty_indexes = [idx for idx, value in enumerate(values, start=1) if not is_blank(value)]
        text_like_indexes = [
            idx
            for idx, value in enumerate(values, start=1)
            if isinstance(value, str) and bool(value.strip())
        ]
        data_rows = 0
        for candidate_row in range(row_index + 1, min(row_count, row_index + 10) + 1):
            if any(not is_blank(value) for value in values_row(cached_rows, candidate_row, column_count)):
                data_rows += 1
        if (
            len(non_empty_indexes) >= min_non_empty_header_cells
            and len(text_like_indexes) >= min_non_empty_header_cells
            and data_rows >= min_data_rows_after_header
        ):
            confidence = min(
                1.0,
                0.45
                + min(0.35, len(text_like_indexes) / max(column_count, 1))
                + min(0.20, data_rows / 10),
            )
            headers = [
                safe_text(values[index - 1]) or f"column_{index}"
                for index in range(1, column_count + 1)
            ]
            if best is None or confidence > best[1]:
                best = (row_index, confidence, headers, non_empty_indexes)

    if best is None:
        return HeaderDetection(
            status=MappingStatus.UNRESOLVED,
            header_row=None,
            confidence=0,
            column_names=[],
            value_columns=[],
            label_columns=[],
            year_columns=[],
            age_columns=[],
            warnings=["No header row met conservative detection thresholds."],
        )

    row_index, confidence, headers, columns = best
    status = (
        MappingStatus.INFERRED
        if confidence >= min_confidence_for_extraction
        else MappingStatus.UNRESOLVED
    )
    warnings = [] if status is MappingStatus.INFERRED else ["Detected header confidence was too low."]
    return HeaderDetection(
        status=status,
        header_row=row_index,
        confidence=confidence,
        column_names=headers,
        value_columns=columns,
        label_columns=[],
        year_columns=[],
        age_columns=[],
        warnings=warnings,
    )


def source_ref_for_file(source_id_prefix: str, relative_path: Path, digest: str) -> SourceRef:
    """Create a source reference for a local workbook."""

    safe_path = "_".join(relative_path.parts).replace(" ", "_")
    return SourceRef(
        source_id=f"{source_id_prefix}:{safe_path}:{digest[:12]}",
        kind=SourceKind.MANUAL_FILE,
        path=relative_path,
        sha256=digest,
        notes="Real local workbook source; contents remain local-only.",
    )


def provenance_for_cell(
    provenance_prefix: str,
    source_ref: SourceRef,
    source_table: SourceTable,
    row_number: int,
    column_number: int,
    header: str,
) -> ProvenanceRecord:
    """Create provenance for one extracted source cell."""

    column_letter = get_column_letter(column_number)
    return ProvenanceRecord(
        provenance_id=(
            f"{provenance_prefix}:{source_table.table_id}:"
            f"{source_table.sheet_name}:{column_letter}{row_number}"
        ),
        source_ref=source_ref,
        claim="Source cell was normalized without semantic reinterpretation.",
        locator=f"{source_table.sheet_name}!{column_letter}{row_number}",
        input_refs=[source_table.table_id],
        notes=f"Original header: {safe_text(header)}",
    )


def record_from_cell(
    *,
    record_prefix: str,
    source_ref: SourceRef,
    source_table: SourceTable,
    row_number: int,
    column_number: int,
    header: str,
    value: Any,
    value_role: str | None = None,
    mapping_override_id: str | None = None,
    geography_labels: dict[str, str] | None = None,
    time_labels: dict[str, str] | None = None,
    age_labels: dict[str, str] | None = None,
) -> NormalizedRecord | None:
    """Build a normalized record from a single nonblank source cell."""

    value_type = infer_value_type(value)
    if value_type is None:
        return None
    normalized_value = normalize_cell_value(value, value_type)
    column_letter = get_column_letter(column_number)
    field_name = normalize_field_name(header, column_number)
    provenance = provenance_for_cell(
        record_prefix,
        source_ref,
        source_table,
        row_number,
        column_number,
        header,
    )
    return NormalizedRecord(
        record_id=(
            f"{record_prefix}:{source_table.table_id}:"
            f"{source_table.sheet_name}:{column_letter}{row_number}"
        ),
        source_table_id=source_table.table_id,
        source_file_path=source_table.workbook_path,
        source_file_hash=source_table.source_file_hash,
        source_sheet=source_table.sheet_name,
        source_row_number=row_number,
        source_column_number=column_number,
        normalized_field_name=field_name,
        normalized_value=normalized_value,
        original_column=column_letter,
        original_header=safe_text(header),
        value_type=value_type,
        value_role=value_role,
        mapping_override_id=mapping_override_id,
        geography_labels=geography_labels or {},
        time_labels=time_labels or {},
        age_labels=age_labels or {},
        provenance=[provenance],
    )


def make_source_table(
    *,
    source_type: str,
    relative_path: Path,
    digest: str,
    sheet: Any,
    detection: HeaderDetection,
    table_role: str,
) -> SourceTable:
    """Create a source table model from a sheet and detection result."""

    return SourceTable(
        table_id=source_table_id(source_type, relative_path, sheet.title),
        workbook_path=relative_path,
        source_file_hash=digest,
        sheet_name=sheet.title,
        row_count=int(sheet.max_row or 0),
        column_count=int(sheet.max_column or 0),
        detected_header_row=detection.header_row,
        column_names=detection.column_names,
        table_role=table_role,
        mapping_status=detection.status,
        warnings=detection.warnings,
    )


def make_mapping(
    *,
    mapping_prefix: str,
    source_table: SourceTable,
    detection: HeaderDetection,
    notes: str | None = None,
) -> ExtractionMapping:
    """Create an extraction mapping from a header detection."""

    return ExtractionMapping(
        mapping_id=f"{mapping_prefix}:{source_table.table_id}",
        source_table_id=source_table.table_id,
        sheet_name=source_table.sheet_name,
        status=detection.status,
        confidence=detection.confidence,
        header_row=detection.header_row,
        value_columns=[] if detection.status is MappingStatus.UNRESOLVED else detection.value_columns,
        label_columns=[] if detection.status is MappingStatus.UNRESOLVED else detection.label_columns,
        year_columns=[] if detection.status is MappingStatus.UNRESOLVED else detection.year_columns,
        age_columns=[] if detection.status is MappingStatus.UNRESOLVED else detection.age_columns,
        notes=notes,
        warnings=detection.warnings,
    )


def write_outputs(
    *,
    root: Path,
    records_path: Path,
    source_tables_path: Path,
    report_json_path: Path,
    report_md_path: Path,
    manifest_path: Path,
    source_tables: list[SourceTable],
    records: list[NormalizedRecord],
    mappings: list[ExtractionMapping],
    manifest: NormalizationManifest,
    report_title: str,
) -> dict[str, str]:
    """Write normalized records, source tables, reports, and manifest."""

    write_jsonl(root / records_path, records)
    write_json(
        root / source_tables_path,
        [table.model_dump(mode="json") for table in source_tables],
    )
    report_payload = {
        "source_tables": [table.model_dump(mode="json") for table in source_tables],
        "mappings": [mapping.model_dump(mode="json") for mapping in mappings],
        "summary": {
            "source_table_count": len(source_tables),
            "normalized_record_count": len(records),
            "unresolved_mapping_count": sum(
                1 for mapping in mappings if mapping.status is MappingStatus.UNRESOLVED
            ),
            "warnings": manifest.warnings,
        },
    }
    write_json(root / report_json_path, report_payload)
    write_markdown_report(
        root / report_md_path,
        report_title,
        source_tables,
        mappings,
        len(records),
        manifest.warnings,
    )
    write_json(root / manifest_path, manifest.model_dump(mode="json"))
    return {
        "normalized_records": str(records_path),
        "source_tables": str(source_tables_path),
        "mapping_report_json": str(report_json_path),
        "mapping_report_markdown": str(report_md_path),
        "manifest": str(manifest_path),
    }
