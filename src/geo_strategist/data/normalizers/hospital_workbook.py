"""Deterministic normalizer for the hospital workbook."""

from __future__ import annotations

import warnings as warning_lib
from pathlib import Path

from openpyxl import load_workbook

from geo_strategist import __version__
from geo_strategist.data.inventory import sha256_file
from geo_strategist.data.normalization import MappingStatus, NormalizationManifest, NormalizationResult, now_utc
from geo_strategist.data.normalizers.common import (
    detect_generic_header,
    load_yaml,
    make_mapping,
    make_source_table,
    read_inspection_profile,
    record_from_cell,
    source_ref_for_file,
    values_row,
    worksheet_values,
    write_outputs,
)
from geo_strategist.data.workbook import locate_hospital_workbook


CONFIG_PATH = Path("configs/source_mappings/hospital_workbook.yaml")


def normalize_hospital_workbook(
    repo_root: str | Path = ".",
    config_path: str | Path = CONFIG_PATH,
) -> NormalizationResult:
    """Normalize clear tabular cells from the hospital workbook.

    This function performs only structural source normalization. It does not
    calculate profitability, derive demand, or generate cash-flow projections.
    """

    root = Path(repo_root).resolve()
    config = load_yaml(config_path)
    outputs = config["outputs"]
    policy = config["extraction_policy"]
    started_at = now_utc()
    workbook_path, source_location, warnings = locate_hospital_workbook(root)
    profile_path = root / config["source"]["inspection_profile"]
    profile = read_inspection_profile(profile_path)
    if profile is None:
        warnings.append("Inspection profile was not found; normalizer inspected workbook directly.")

    source_tables = []
    mappings = []
    records = []
    input_files = []

    if workbook_path is not None:
        relative_path = workbook_path.relative_to(root)
        digest = sha256_file(workbook_path)
        source_ref = source_ref_for_file("hospital_workbook", relative_path, digest)
        input_files.append(source_ref)

        with warning_lib.catch_warnings(record=True) as caught_warnings:
            warning_lib.simplefilter("always", UserWarning)
            workbook = load_workbook(workbook_path, read_only=True, data_only=False)
            for sheet in workbook.worksheets:
                rows = worksheet_values(sheet)
                detection = detect_generic_header(
                    sheet,
                    max_header_scan_rows=int(policy["max_header_scan_rows"]),
                    min_non_empty_header_cells=int(policy["min_non_empty_header_cells"]),
                    min_data_rows_after_header=int(policy["min_data_rows_after_header"]),
                    min_confidence_for_extraction=float(policy["min_confidence_for_extraction"]),
                    rows=rows,
                )
                source_table = make_source_table(
                    source_type="hospital_workbook",
                    relative_path=relative_path,
                    digest=digest,
                    sheet=sheet,
                    detection=detection,
                    table_role="hospital_workbook_sheet",
                )
                mapping = make_mapping(
                    mapping_prefix="hospital_mapping",
                    source_table=source_table,
                    detection=detection,
                    notes=(
                        "Inferred from workbook structure only; business meaning was not interpreted."
                    ),
                )
                source_tables.append(source_table)
                mappings.append(mapping)

                if mapping.status is MappingStatus.UNRESOLVED or mapping.header_row is None:
                    continue

                for row_index in range(mapping.header_row + 1, int(sheet.max_row or 0) + 1):
                    values = values_row(rows, row_index, int(sheet.max_column or 0))
                    for column_number in mapping.value_columns:
                        if column_number > len(values):
                            continue
                        header = source_table.column_names[column_number - 1]
                        record = record_from_cell(
                            record_prefix="hospital_record",
                            source_ref=source_ref,
                            source_table=source_table,
                            row_number=row_index,
                            column_number=column_number,
                            header=header,
                            value=values[column_number - 1],
                        )
                        if record is not None:
                            records.append(record)
            workbook.close()
        warnings.extend(sorted({str(warning.message) for warning in caught_warnings}))
    else:
        warnings.extend(["Hospital workbook was not normalized because no input file was found."])

    if source_location == "legacy_fallback":
        warnings.append("Used legacy data/manual fallback; prefer canonical .data/manual input.")

    unresolved_count = sum(1 for mapping in mappings if mapping.status is MappingStatus.UNRESOLVED)
    output_files = [
        Path(outputs["normalized_records"]),
        Path(outputs["source_tables"]),
        Path(outputs["mapping_report_json"]),
        Path(outputs["mapping_report_markdown"]),
        Path(outputs["manifest"]),
    ]
    manifest = NormalizationManifest(
        run_id=f"hospital_workbook_normalization:{started_at.isoformat()}",
        source_type="hospital_workbook",
        started_at=started_at,
        ended_at=now_utc(),
        input_files=input_files,
        output_files=output_files,
        source_table_count=len(source_tables),
        normalized_record_count=len(records),
        unresolved_mapping_count=unresolved_count,
        warnings=warnings,
        code_version=__version__,
    )
    output_paths = write_outputs(
        root=root,
        records_path=Path(outputs["normalized_records"]),
        source_tables_path=Path(outputs["source_tables"]),
        report_json_path=Path(outputs["mapping_report_json"]),
        report_md_path=Path(outputs["mapping_report_markdown"]),
        manifest_path=Path(outputs["manifest"]),
        source_tables=source_tables,
        records=records,
        mappings=mappings,
        manifest=manifest,
        report_title="Hospital Workbook Normalization Mapping Report",
    )
    return NormalizationResult(
        found=workbook_path is not None,
        source_type="hospital_workbook",
        source_tables=source_tables,
        normalized_records=records,
        mappings=mappings,
        manifest=manifest,
        output_paths=output_paths,
        warnings=warnings,
    )
