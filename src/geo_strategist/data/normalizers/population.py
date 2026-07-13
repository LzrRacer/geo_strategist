"""Deterministic normalizer for local population workbooks."""

from __future__ import annotations

import warnings as warning_lib
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from geo_strategist import __version__
from geo_strategist.data.inventory import sha256_file
from geo_strategist.data.normalization import (
    MappingStatus,
    NormalizationManifest,
    NormalizationResult,
    now_utc,
    safe_text,
)
from geo_strategist.data.normalizers.common import (
    HeaderDetection,
    is_blank,
    load_yaml,
    make_mapping,
    make_source_table,
    read_inspection_profile,
    record_from_cell,
    source_ref_for_file,
    values_row,
    worksheet_values,
    write_outputs,
)
from geo_strategist.data.population import locate_population_dir


CONFIG_PATH = Path("configs/source_mappings/population_workbooks.yaml")


def _match_categories(text: str, keyword_groups: dict[str, list[str]]) -> list[str]:
    lowered = text.lower()
    categories: list[str] = []
    for category, keywords in keyword_groups.items():
        if any(keyword.lower() in lowered for keyword in keywords):
            categories.append(category)
    return categories


def _detect_population_header(
    sheet: Any,
    *,
    rows: list[list[Any]],
    keyword_groups: dict[str, list[str]],
    max_header_scan_rows: int,
    min_non_empty_header_cells: int,
    min_data_rows_after_header: int,
    min_confidence_for_extraction: float,
) -> HeaderDetection:
    row_count = len(rows)
    column_count = int(sheet.max_column or 0)
    best: tuple[int, float, list[str], dict[str, list[int]], list[str]] | None = None

    for row_index in range(1, min(row_count, max_header_scan_rows) + 1):
        values = values_row(rows, row_index, column_count)
        non_empty = [idx for idx, value in enumerate(values, start=1) if not is_blank(value)]
        if len(non_empty) < min_non_empty_header_cells:
            continue

        category_columns: dict[str, list[int]] = {
            "prefecture": [],
            "municipality": [],
            "age": [],
            "year": [],
            "population": [],
        }
        headers = []
        for column_index in range(1, column_count + 1):
            header = safe_text(values[column_index - 1] if column_index - 1 < len(values) else None)
            headers.append(header or f"column_{column_index}")
            for category in _match_categories(header, keyword_groups):
                category_columns.setdefault(category, []).append(column_index)

        matched_categories = {
            category for category, columns in category_columns.items() if columns
        }
        data_rows = 0
        for candidate_row in range(row_index + 1, min(row_count, row_index + 10) + 1):
            if any(not is_blank(value) for value in values_row(rows, candidate_row, column_count)):
                data_rows += 1

        has_required = "population" in matched_categories and bool(
            matched_categories & {"prefecture", "municipality", "age", "year"}
        )
        if not has_required or data_rows < min_data_rows_after_header:
            continue

        population_column_count = len(category_columns.get("population", []))
        confidence = min(
            1.0,
            0.40
            + 0.12 * len(matched_categories)
            + min(0.20, data_rows / 10),
        )
        warnings = []
        if population_column_count != 1:
            warnings.append(
                "Expected exactly one population value column; sheet is ambiguous."
            )
        if "municipality" not in matched_categories:
            warnings.append("Municipality column was not detected.")
        if "age" not in matched_categories:
            warnings.append("Age column was not detected.")
        if "year" not in matched_categories:
            warnings.append("Year column was not detected.")

        if best is None or confidence > best[1]:
            best = (row_index, confidence, headers, category_columns, warnings)

    if best is None:
        return HeaderDetection(
            status=MappingStatus.UNRESOLVED,
            header_row=None,
            confidence=0,
            column_names=[],
            value_columns=[],
            label_columns=[],
            year_columns=[],
            age_columns=[],
            warnings=["No population header row met conservative keyword thresholds."],
        )

    row_index, confidence, headers, category_columns, warnings = best
    has_single_population_column = len(category_columns.get("population", [])) == 1
    status = (
        MappingStatus.INFERRED
        if confidence >= min_confidence_for_extraction and has_single_population_column
        else MappingStatus.UNRESOLVED
    )
    if status is MappingStatus.UNRESOLVED:
        warnings.append("Detected population header confidence was too low.")
    return HeaderDetection(
        status=status,
        header_row=row_index,
        confidence=confidence,
        column_names=headers,
        value_columns=category_columns.get("population", []),
        label_columns=[
            *category_columns.get("prefecture", []),
            *category_columns.get("municipality", []),
        ],
        year_columns=category_columns.get("year", []),
        age_columns=category_columns.get("age", []),
        warnings=warnings,
    )


def _normalise_year_map(raw: dict[Any, Any] | None) -> dict[int, str]:
    if not raw:
        return {}
    return {int(key): str(value).replace("年", "") for key, value in raw.items()}


def _matching_override(
    overrides: list[dict[str, Any]],
    relative_path: Path,
    sheet_name: str,
) -> dict[str, Any] | None:
    path_text = str(relative_path)
    for override in overrides:
        if override.get("workbook_path") == path_text and override.get("sheet_name") == sheet_name:
            return override
    return None


def _is_verified_override(override: dict[str, Any] | None) -> bool:
    if not override:
        return False
    return (
        override.get("status") == "manual_verified"
        and override.get("population_value_type") in {"population_count", "population_rate"}
        and override.get("time_label_source") != "unresolved"
        and bool(_normalise_year_map(override.get("year_by_column")))
    )


def _filled_headers(row: list[Any], column_count: int) -> list[str]:
    headers: list[str] = []
    last_header = ""
    for column_index in range(1, column_count + 1):
        text = safe_text(row[column_index - 1] if column_index - 1 < len(row) else None)
        if text:
            last_header = text
        headers.append(text or last_header or f"column_{column_index}")
    return headers


def _detection_from_override(
    override: dict[str, Any],
    rows: list[list[Any]],
    column_count: int,
) -> HeaderDetection:
    header_row = int(override["header_row"])
    value_columns = [int(column) for column in override.get("value_columns", [])]
    geography_columns = [int(column) for column in override.get("geography_columns", [])]
    age_columns = [int(column) for column in override.get("age_columns", [])]
    year_columns = list(_normalise_year_map(override.get("year_by_column")).keys())
    max_column = max([*value_columns, *geography_columns, *age_columns, *year_columns, 0])
    warnings: list[str] = []
    if header_row < 1 or header_row > len(rows):
        warnings.append("Manual override header row is outside worksheet dimensions.")
    if max_column > column_count:
        warnings.append("Manual override references columns outside worksheet dimensions.")
    if warnings:
        return HeaderDetection(
            status=MappingStatus.UNRESOLVED,
            header_row=None,
            confidence=0,
            column_names=[],
            value_columns=[],
            label_columns=[],
            year_columns=[],
            age_columns=[],
            warnings=warnings,
        )
    headers = _filled_headers(values_row(rows, header_row, column_count), column_count)
    return HeaderDetection(
        status=MappingStatus.MANUAL_VERIFIED,
        header_row=header_row,
        confidence=1,
        column_names=headers,
        value_columns=value_columns,
        label_columns=geography_columns,
        year_columns=year_columns,
        age_columns=age_columns,
        warnings=[],
    )


def _label_map(values: list[Any], headers: list[str], columns: list[int]) -> dict[str, str]:
    labels: dict[str, str] = {}
    for column_number in columns:
        if column_number > len(values):
            continue
        value = values[column_number - 1]
        if is_blank(value):
            continue
        labels[headers[column_number - 1]] = safe_text(value)
    return labels


def _override_age_labels(override: dict[str, Any]) -> dict[str, str]:
    age_label = override.get("age_label")
    return {"age_group": str(age_label)} if age_label else {}


def _override_time_labels(override: dict[str, Any], column_number: int) -> dict[str, str]:
    year = _normalise_year_map(override.get("year_by_column")).get(column_number)
    return {"year": year} if year else {}


def normalize_population_data(
    repo_root: str | Path = ".",
    config_path: str | Path = CONFIG_PATH,
) -> NormalizationResult:
    """Normalize clearly mapped population workbook cells.

    This function does not reshape final population tables or calculate demand.
    """

    root = Path(repo_root).resolve()
    config = load_yaml(config_path)
    outputs = config["outputs"]
    policy = config["extraction_policy"]
    started_at = now_utc()
    source_dir, source_location, warnings = locate_population_dir(root)
    profile_path = root / config["source"]["inspection_profile"]
    profile = read_inspection_profile(profile_path)
    if profile is None:
        warnings.append("Inspection profile was not found; normalizer inspected workbooks directly.")

    source_tables = []
    mappings = []
    records = []
    input_files = []
    manual_overrides = config.get("manual_overrides", [])

    if source_dir is not None:
        workbook_paths = sorted(path for path in source_dir.glob("*.xlsx") if path.is_file())
        if not workbook_paths:
            warnings.append("No .xlsx population workbooks were found.")

        for workbook_path in workbook_paths:
            relative_path = workbook_path.relative_to(root)
            digest = sha256_file(workbook_path)
            source_ref = source_ref_for_file("population_workbook", relative_path, digest)
            input_files.append(source_ref)

            with warning_lib.catch_warnings(record=True) as caught_warnings:
                warning_lib.simplefilter("always", UserWarning)
                workbook = load_workbook(workbook_path, read_only=True, data_only=False)
                for sheet in workbook.worksheets:
                    rows = worksheet_values(sheet)
                    override = _matching_override(manual_overrides, relative_path, sheet.title)
                    is_manual_override = _is_verified_override(override)
                    if is_manual_override and override is not None:
                        detection = _detection_from_override(
                            override,
                            rows,
                            int(sheet.max_column or 0),
                        )
                        mapping_notes = (
                            f"Manual override {override['id']}: "
                            f"{override.get('evidence_text', '')}"
                        )
                    else:
                        detection = _detect_population_header(
                            sheet,
                            rows=rows,
                            keyword_groups=config["keyword_groups"],
                            max_header_scan_rows=int(policy["max_header_scan_rows"]),
                            min_non_empty_header_cells=int(policy["min_non_empty_header_cells"]),
                            min_data_rows_after_header=int(policy["min_data_rows_after_header"]),
                            min_confidence_for_extraction=float(policy["min_confidence_for_extraction"]),
                        )
                        mapping_notes = "Inferred using conservative population keyword matching."
                    source_table = make_source_table(
                        source_type="population",
                        relative_path=relative_path,
                        digest=digest,
                        sheet=sheet,
                        detection=detection,
                        table_role="population_workbook_sheet",
                    )
                    mapping = make_mapping(
                        mapping_prefix="population_mapping",
                        source_table=source_table,
                        detection=detection,
                        notes=mapping_notes,
                    )
                    source_tables.append(source_table)
                    mappings.append(mapping)

                    if mapping.status is MappingStatus.UNRESOLVED or mapping.header_row is None:
                        continue

                    headers = source_table.column_names
                    start_row = mapping.header_row + 1
                    if is_manual_override and override is not None and override.get("time_label_row"):
                        start_row = max(start_row, int(override["time_label_row"]) + 1)
                    for row_index in range(start_row, int(sheet.max_row or 0) + 1):
                        values = values_row(rows, row_index, int(sheet.max_column or 0))
                        geography_labels = _label_map(values, headers, mapping.label_columns)
                        base_time_labels = (
                            {}
                            if is_manual_override and override is not None
                            else _label_map(values, headers, mapping.year_columns)
                        )
                        age_labels = (
                            _override_age_labels(override)
                            if is_manual_override and override is not None
                            else _label_map(values, headers, mapping.age_columns)
                        )
                        for column_number in mapping.value_columns:
                            if column_number > len(values):
                                continue
                            header = headers[column_number - 1]
                            time_labels = base_time_labels
                            value_role = None
                            override_id = None
                            if is_manual_override and override is not None:
                                time_labels = _override_time_labels(override, column_number)
                                year = time_labels.get("year")
                                if year:
                                    header = f"{header} {year}年"
                                value_role = override.get("population_value_type")
                                override_id = override.get("id")
                            record = record_from_cell(
                                record_prefix="population_record",
                                source_ref=source_ref,
                                source_table=source_table,
                                row_number=row_index,
                                column_number=column_number,
                                header=header,
                                value=values[column_number - 1],
                                value_role=value_role,
                                mapping_override_id=override_id,
                                geography_labels=geography_labels,
                                time_labels=time_labels,
                                age_labels=age_labels,
                            )
                            if record is not None:
                                records.append(record)
                workbook.close()
            warnings.extend(sorted({str(warning.message) for warning in caught_warnings}))
    else:
        warnings.append("Population workbooks were not normalized because no input directory was found.")

    if source_location == "legacy_fallback":
        warnings.append("Used legacy data/manual fallback; prefer canonical .data/manual input.")

    unresolved_count = sum(1 for mapping in mappings if mapping.status is MappingStatus.UNRESOLVED)
    output_files = [
        Path(outputs["normalized_records"]),
        Path(outputs["source_tables"]),
        Path(outputs["mapping_report_json"]),
        Path(outputs["mapping_report_markdown"]),
        Path(outputs["manifest"]),
    ]
    manifest = NormalizationManifest(
        run_id=f"population_normalization:{started_at.isoformat()}",
        source_type="population",
        started_at=started_at,
        ended_at=now_utc(),
        input_files=input_files,
        output_files=output_files,
        source_table_count=len(source_tables),
        normalized_record_count=len(records),
        unresolved_mapping_count=unresolved_count,
        warnings=warnings,
        code_version=__version__,
    )
    output_paths = write_outputs(
        root=root,
        records_path=Path(outputs["normalized_records"]),
        source_tables_path=Path(outputs["source_tables"]),
        report_json_path=Path(outputs["mapping_report_json"]),
        report_md_path=Path(outputs["mapping_report_markdown"]),
        manifest_path=Path(outputs["manifest"]),
        source_tables=source_tables,
        records=records,
        mappings=mappings,
        manifest=manifest,
        report_title="Population Normalization Mapping Report",
    )
    return NormalizationResult(
        found=source_dir is not None and bool(input_files),
        source_type="population",
        source_tables=source_tables,
        normalized_records=records,
        mappings=mappings,
        manifest=manifest,
        output_paths=output_paths,
        warnings=warnings,
    )
