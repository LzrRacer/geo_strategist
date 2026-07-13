"""Population workbook inspection utilities.

This module profiles structure and conservative label hints only. It does not
reshape population tables, invent age groups, or calculate demand.
"""

from __future__ import annotations

import json
import warnings as warning_lib
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, ConfigDict, Field

from geo_strategist.data.inventory import sha256_file


CANONICAL_POPULATION_DIR = Path(".data/manual/population_data")
LEGACY_POPULATION_DIR = Path("data/manual/population_data")
DEFAULT_POPULATION_PROFILE_JSON = Path(".cache/inspection/population_data_profile.json")
DEFAULT_POPULATION_PROFILE_MD = Path(".cache/inspection/population_data_profile.md")

HEADER_SCAN_ROWS = 30

KEYWORDS = {
    "prefecture": ("prefecture", "都道府県", "都道府県名"),
    "municipality": ("municipality", "市区町村", "市町村", "市区町村名", "市町村名"),
    "age": ("age", "年齢", "歳", "年齢階級"),
    "year": ("year", "年度", "年次", "調査年", "基準年"),
    "population": ("population", "人口", "総人口"),
}


class HeaderCandidate(BaseModel):
    """Likely header row with matched label categories."""

    model_config = ConfigDict(extra="forbid")

    row_index: int = Field(ge=1)
    matched_categories: list[str] = Field(default_factory=list)


class LabelColumnCandidate(BaseModel):
    """Likely label column from conservative keyword matching."""

    model_config = ConfigDict(extra="forbid")

    category: str
    column_index: int = Field(ge=1)
    column_letter: str
    header_text: str
    matched_keyword: str


class PopulationSheetProfile(BaseModel):
    """Structural profile for one population worksheet."""

    model_config = ConfigDict(extra="forbid")

    name: str
    row_count: int = Field(ge=0)
    column_count: int = Field(ge=0)
    likely_header_rows: list[HeaderCandidate] = Field(default_factory=list)
    likely_label_columns: list[LabelColumnCandidate] = Field(default_factory=list)


class PopulationFileProfile(BaseModel):
    """Structural profile for one population workbook."""

    model_config = ConfigDict(extra="forbid")

    path: str
    sha256: str = Field(pattern=r"^[a-fA-F0-9]{64}$")
    sheet_names: list[str] = Field(default_factory=list)
    sheets: list[PopulationSheetProfile] = Field(default_factory=list)


class PopulationProfile(BaseModel):
    """Population inspection result."""

    model_config = ConfigDict(extra="forbid")

    found: bool
    source_directory: str | None = None
    files: list[PopulationFileProfile] = Field(default_factory=list)
    warnings: list[str] = Field(default_factory=list)
    output_json: str | None = None
    output_markdown: str | None = None


def locate_population_dir(repo_root: str | Path = ".") -> tuple[Path | None, str | None, list[str]]:
    """Locate population workbook directory, falling back to legacy data/manual."""

    root = Path(repo_root).resolve()
    canonical = root / CANONICAL_POPULATION_DIR
    legacy = root / LEGACY_POPULATION_DIR
    if canonical.exists():
        return canonical, "canonical", []
    if legacy.exists():
        return legacy, "legacy_fallback", [
            "Using legacy data/manual population fallback; prefer .data/manual."
        ]
    return None, None, ["Population workbook directory was not found."]


def _safe_text(value: Any, limit: int = 80) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()[:limit]


def _match_categories(text: str) -> list[tuple[str, str]]:
    lowered = text.lower()
    matches: list[tuple[str, str]] = []
    for category, keywords in KEYWORDS.items():
        for keyword in keywords:
            if keyword.lower() in lowered:
                matches.append((category, keyword))
                break
    return matches


def _profile_sheet(sheet: Any) -> PopulationSheetProfile:
    row_count = int(sheet.max_row or 0)
    column_count = int(sheet.max_column or 0)
    header_candidates: list[HeaderCandidate] = []
    column_candidates: dict[tuple[str, int], LabelColumnCandidate] = {}

    for row_index, row in enumerate(
        sheet.iter_rows(
            min_row=1,
            max_row=min(row_count, HEADER_SCAN_ROWS),
            values_only=True,
        ),
        start=1,
    ):
        row_categories: set[str] = set()
        for column_index, value in enumerate(row[:column_count], start=1):
            text = _safe_text(value)
            if not text:
                continue
            for category, keyword in _match_categories(text):
                row_categories.add(category)
                key = (category, column_index)
                column_candidates.setdefault(
                    key,
                    LabelColumnCandidate(
                        category=category,
                        column_index=column_index,
                        column_letter=get_column_letter(column_index),
                        header_text=text,
                        matched_keyword=keyword,
                    ),
                )
        if row_categories:
            header_candidates.append(
                HeaderCandidate(
                    row_index=row_index,
                    matched_categories=sorted(row_categories),
                )
            )

    return PopulationSheetProfile(
        name=sheet.title,
        row_count=row_count,
        column_count=column_count,
        likely_header_rows=header_candidates[:10],
        likely_label_columns=sorted(
            column_candidates.values(),
            key=lambda candidate: (candidate.category, candidate.column_index),
        ),
    )


def _markdown_summary(profile: PopulationProfile) -> str:
    lines = ["# Population Data Profile", ""]
    if not profile.found:
        lines.extend(["Population workbook files not found.", ""])
        for warning in profile.warnings:
            lines.append(f"- Warning: {warning}")
        return "\n".join(lines).rstrip() + "\n"

    lines.extend(
        [
            f"- Source directory: `{profile.source_directory}`",
            f"- Workbook files: {len(profile.files)}",
            "",
            "| File | Sheets | SHA256 |",
            "| --- | ---: | --- |",
        ]
    )
    for file_profile in profile.files:
        lines.append(
            f"| {file_profile.path} | {len(file_profile.sheet_names)} | "
            f"`{file_profile.sha256}` |"
        )
    lines.extend(["", "## Sheets", ""])
    lines.append("| File | Sheet | Rows | Columns | Header hints | Label hints |")
    lines.append("| --- | --- | ---: | ---: | ---: | ---: |")
    for file_profile in profile.files:
        for sheet in file_profile.sheets:
            lines.append(
                f"| {file_profile.path} | {sheet.name} | {sheet.row_count} | "
                f"{sheet.column_count} | {len(sheet.likely_header_rows)} | "
                f"{len(sheet.likely_label_columns)} |"
            )
    if profile.warnings:
        lines.extend(["", "## Warnings"])
        lines.extend(f"- {warning}" for warning in profile.warnings)
    return "\n".join(lines).rstrip() + "\n"


def inspect_population_data(
    repo_root: str | Path = ".",
    output_json: str | Path = DEFAULT_POPULATION_PROFILE_JSON,
    output_markdown: str | Path = DEFAULT_POPULATION_PROFILE_MD,
) -> PopulationProfile:
    """Inspect population workbook structures and write ignored profile files."""

    root = Path(repo_root).resolve()
    source_dir, source_location, warnings = locate_population_dir(root)

    file_profiles: list[PopulationFileProfile] = []
    if source_dir is not None:
        workbook_paths = sorted(path for path in source_dir.glob("*.xlsx") if path.is_file())
        if not workbook_paths:
            warnings.append("No .xlsx population workbooks were found in the source directory.")
        for path in workbook_paths:
            with warning_lib.catch_warnings(record=True) as caught_warnings:
                warning_lib.simplefilter("always", UserWarning)
                workbook = load_workbook(path, read_only=True, data_only=False)
                sheet_names = list(workbook.sheetnames)
                sheets = [_profile_sheet(workbook[sheet_name]) for sheet_name in sheet_names]
                workbook.close()
            warnings.extend(sorted({str(warning.message) for warning in caught_warnings}))
            file_profiles.append(
                PopulationFileProfile(
                    path=str(path.relative_to(root)),
                    sha256=sha256_file(path),
                    sheet_names=sheet_names,
                    sheets=sheets,
                )
            )

    profile = PopulationProfile(
        found=bool(file_profiles),
        source_directory=str(source_dir.relative_to(root)) if source_dir else None,
        files=file_profiles,
        warnings=warnings,
    )
    if source_location == "legacy_fallback" and profile.source_directory:
        profile.warnings.append("Population profile used legacy fallback data/manual.")

    json_path = root / output_json
    markdown_path = root / output_markdown
    json_path.parent.mkdir(parents=True, exist_ok=True)
    markdown_path.parent.mkdir(parents=True, exist_ok=True)
    profile.output_json = str(json_path.relative_to(root))
    profile.output_markdown = str(markdown_path.relative_to(root))
    json_path.write_text(
        json.dumps(profile.model_dump(mode="json"), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    markdown_path.write_text(_markdown_summary(profile), encoding="utf-8")
    return profile
