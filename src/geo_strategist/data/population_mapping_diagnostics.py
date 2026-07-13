"""Diagnostics for explicit population mapping resolution."""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field

from geo_strategist.data.normalization import safe_text
from geo_strategist.data.normalizers.common import is_blank, values_row, worksheet_values


PROFILE_PATH = Path(".cache/inspection/population_data_profile.json")
REPORT_PATH = Path(".cache/normalization/population_mapping_report.json")
MANUAL_CANDIDATES_PATH = Path(".cache/review/manual_mapping_candidates.yaml")
QUALITY_ISSUES_PATH = Path(".cache/views/population_quality_issues.jsonl")
OUTPUT_JSON = Path(".cache/review/population_mapping_diagnostics.json")
OUTPUT_MD = Path(".cache/review/population_mapping_diagnostics.md")

TIME_PATTERNS = [
    re.compile(r"(?:19|20)\d{2}"),
    re.compile(r"令和\s*\d+\s*年"),
    re.compile(r"平成\s*\d+\s*年"),
    re.compile(r"昭和\s*\d+\s*年"),
]


class PopulationMappingDiagnosticsResult(BaseModel):
    """Population mapping diagnostics summary."""

    model_config = ConfigDict(extra="forbid")

    source_table_count: int = 0
    diagnostics_count: int = 0
    unresolved_mapping_count: int = 0
    quarantined_issue_count: int = 0
    output_paths: dict[str, str] = Field(default_factory=dict)


def _load_json(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def _load_yaml(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    return yaml.safe_load(path.read_text(encoding="utf-8"))


def _quality_issue_counts(path: Path) -> dict[tuple[str, str], Counter[str]]:
    counts: dict[tuple[str, str], Counter[str]] = defaultdict(Counter)
    if not path.exists():
        return counts
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        payload = json.loads(line)
        key = (payload.get("source_file") or "", payload.get("sheet") or "")
        counts[key][payload["issue_type"]] += 1
    return counts


def _strings_from_row(row: list[Any], limit: int = 20) -> list[str]:
    strings = []
    for value in row:
        text = safe_text(value, limit=120)
        if text and text not in strings:
            strings.append(text)
        if len(strings) >= limit:
            break
    return strings


def _time_evidence(*texts: str) -> list[str]:
    evidence = []
    for text in texts:
        for pattern in TIME_PATTERNS:
            for match in pattern.finditer(text):
                match_text = match.group(0)
                if match_text and match_text not in evidence:
                    evidence.append(match_text)
    return evidence


def _caption_and_time_evidence(
    root: Path,
    file_path: str,
    sheet_name: str,
    detected_header_row: int | None,
) -> tuple[list[str], list[str]]:
    path = root / file_path
    if not path.exists():
        return [], _time_evidence(file_path, sheet_name)

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        sheet = workbook[sheet_name]
        rows = worksheet_values(sheet)
        captions: list[str] = []
        if detected_header_row is not None:
            max_evidence_row = min(len(rows), detected_header_row + 1)
        else:
            max_evidence_row = min(len(rows), 3)
        for row_index in range(1, max_evidence_row + 1):
            for text in _strings_from_row(values_row(rows, row_index, int(sheet.max_column or 0))):
                if text not in captions:
                    captions.append(text)
                if len(captions) >= 20:
                    break
            if len(captions) >= 20:
                break
        evidence = _time_evidence(file_path, sheet_name, *captions)
        return captions, evidence
    finally:
        workbook.close()


def _category_columns(headers: list[str]) -> dict[str, list[dict[str, Any]]]:
    categories = {
        "geography": ("都道府県", "市区町村", "市町村", "prefecture", "municipality"),
        "age": ("年齢", "歳", "年齢階級", "age"),
        "population_like": ("人口", "総人口", "population"),
    }
    found: dict[str, list[dict[str, Any]]] = {key: [] for key in categories}
    for index, header in enumerate(headers, start=1):
        lowered = header.lower()
        for category, needles in categories.items():
            if any(needle.lower() in lowered for needle in needles):
                found[category].append({"column": index, "header": header})
    return found


def diagnose_population_mappings(repo_root: str | Path = ".") -> PopulationMappingDiagnosticsResult:
    """Generate population mapping diagnostics without editing configs."""

    root = Path(repo_root).resolve()
    profile = _load_json(root / PROFILE_PATH) or {}
    report = _load_json(root / REPORT_PATH) or {"source_tables": [], "mappings": []}
    manual_candidates = _load_yaml(root / MANUAL_CANDIDATES_PATH) or {}
    issue_counts = _quality_issue_counts(root / QUALITY_ISSUES_PATH)
    mappings = {mapping["source_table_id"]: mapping for mapping in report.get("mappings", [])}
    diagnostics = []

    candidate_ids = {
        candidate.get("source_table_id")
        for candidate in manual_candidates.get("unresolved_source_tables", [])
    }
    for table in report.get("source_tables", []):
        mapping = mappings.get(table["table_id"], {})
        key = (table.get("workbook_path") or "", table.get("sheet_name") or "")
        counts = dict(issue_counts.get(key, Counter()))
        include = (
            mapping.get("status") == "unresolved"
            or counts
            or table["table_id"] in candidate_ids
        )
        if not include:
            continue
        captions, time_evidence = _caption_and_time_evidence(
            root,
            table.get("workbook_path") or "",
            table.get("sheet_name") or "",
            table.get("detected_header_row"),
        )
        categories = _category_columns(table.get("column_names", []))
        diagnostics.append(
            {
                "source_table_id": table["table_id"],
                "file_path": table.get("workbook_path"),
                "sheet_name": table.get("sheet_name"),
                "dimensions": {
                    "rows": table.get("row_count"),
                    "columns": table.get("column_count"),
                },
                "detected_header_row": table.get("detected_header_row"),
                "detected_headers": table.get("column_names", [])[:20],
                "mapping_status": mapping.get("status"),
                "issue_counts": counts,
                "candidate_geography_columns": categories["geography"],
                "candidate_age_columns": categories["age"],
                "candidate_population_like_columns": categories["population_like"],
                "candidate_time_evidence": {
                    "workbook_name": _time_evidence(Path(table.get("workbook_path") or "").name),
                    "sheet_name": _time_evidence(table.get("sheet_name") or ""),
                    "headers": _time_evidence(*table.get("column_names", [])),
                    "captions": time_evidence,
                },
                "header_caption_evidence": captions[:20],
            }
        )

    payload = {
        "profile_source_files": [item.get("path") for item in profile.get("files", [])],
        "diagnostics": diagnostics,
        "summary": {
            "source_table_count": len(report.get("source_tables", [])),
            "diagnostics_count": len(diagnostics),
            "unresolved_mapping_count": sum(
                1 for mapping in report.get("mappings", []) if mapping.get("status") == "unresolved"
            ),
            "quarantined_issue_count": sum(sum(counter.values()) for counter in issue_counts.values()),
        },
    }
    for path, content in [
        (root / OUTPUT_JSON, json.dumps(payload, ensure_ascii=False, indent=2) + "\n"),
        (root / OUTPUT_MD, _markdown(payload)),
    ]:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(content, encoding="utf-8")

    return PopulationMappingDiagnosticsResult(
        source_table_count=payload["summary"]["source_table_count"],
        diagnostics_count=payload["summary"]["diagnostics_count"],
        unresolved_mapping_count=payload["summary"]["unresolved_mapping_count"],
        quarantined_issue_count=payload["summary"]["quarantined_issue_count"],
        output_paths={
            "diagnostics_json": str(OUTPUT_JSON),
            "diagnostics_markdown": str(OUTPUT_MD),
        },
    )


def _markdown(payload: dict[str, Any]) -> str:
    lines = [
        "# Population Mapping Diagnostics",
        "",
        f"- Source tables: {payload['summary']['source_table_count']}",
        f"- Diagnostics: {payload['summary']['diagnostics_count']}",
        f"- Unresolved mappings: {payload['summary']['unresolved_mapping_count']}",
        f"- Quarantined issues: {payload['summary']['quarantined_issue_count']}",
        "",
        "| File | Sheet | Status | Issues | Time Evidence |",
        "| --- | --- | --- | --- | --- |",
    ]
    for item in payload["diagnostics"]:
        evidence = item["candidate_time_evidence"]
        evidence_values = [
            *evidence.get("workbook_name", []),
            *evidence.get("sheet_name", []),
            *evidence.get("headers", []),
            *evidence.get("captions", []),
        ]
        lines.append(
            f"| {item['file_path']} | {item['sheet_name']} | {item['mapping_status']} | "
            f"{item['issue_counts']} | {', '.join(evidence_values[:10])} |"
        )
    return "\n".join(lines).rstrip() + "\n"
