"""Hospital workbook inspection utilities.

This module profiles workbook structure only. It does not modify workbooks,
generate candidate sites, or calculate cash-flow projections.
"""

from __future__ import annotations

import json
import warnings as warning_lib
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, ConfigDict, Field

from geo_strategist.data.inventory import sha256_file


HOSPITAL_WORKBOOK_NAME = "hospital_rough_cf_payback_model_tokyo_aichi_osaka_beds_updated.xlsx"
CANONICAL_WORKBOOK_PATH = Path(".data/manual/hospital_cf_workbook") / HOSPITAL_WORKBOOK_NAME
LEGACY_WORKBOOK_PATH = Path("data/manual/hospital_cf_workbook") / HOSPITAL_WORKBOOK_NAME
DEFAULT_WORKBOOK_PROFILE_JSON = Path(".cache/inspection/hospital_workbook_profile.json")
DEFAULT_WORKBOOK_PROFILE_MD = Path(".cache/inspection/hospital_workbook_profile.md")


class DetectedColumn(BaseModel):
    """Detected column metadata without storing cell contents."""

    model_config = ConfigDict(extra="forbid")

    column_index: int = Field(ge=1)
    column_letter: str
    name: str
    count: int = Field(ge=0)


class SheetProfile(BaseModel):
    """Structural profile for one worksheet."""

    model_config = ConfigDict(extra="forbid")

    name: str
    row_count: int = Field(ge=0)
    column_count: int = Field(ge=0)
    column_names: list[str] = Field(default_factory=list)
    detected_numeric_columns: list[DetectedColumn] = Field(default_factory=list)
    detected_date_like_columns: list[DetectedColumn] = Field(default_factory=list)
    non_empty_cell_count: int = Field(ge=0)


class WorkbookProfile(BaseModel):
    """Workbook inspection result."""

    model_config = ConfigDict(extra="forbid")

    found: bool
    path: str | None = None
    source_location: str | None = None
    sha256: str | None = Field(default=None, pattern=r"^[a-fA-F0-9]{64}$")
    sheet_names: list[str] = Field(default_factory=list)
    sheets: list[SheetProfile] = Field(default_factory=list)
    warnings: list[str] = Field(default_factory=list)
    output_json: str | None = None
    output_markdown: str | None = None


def locate_hospital_workbook(repo_root: str | Path = ".") -> tuple[Path | None, str | None, list[str]]:
    """Locate the canonical hospital workbook, falling back to legacy data/manual."""

    root = Path(repo_root).resolve()
    canonical = root / CANONICAL_WORKBOOK_PATH
    legacy = root / LEGACY_WORKBOOK_PATH
    if canonical.exists():
        return canonical, "canonical", []
    if legacy.exists():
        return legacy, "legacy_fallback", [
            "Using legacy data/manual hospital workbook fallback; prefer .data/manual."
        ]
    return None, None, ["Hospital workbook was not found."]


def _safe_text(value: Any, limit: int = 80) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return text[:limit]


def _is_numeric(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _is_date_like(value: Any) -> bool:
    return isinstance(value, (date, datetime))


def _column_name(header_value: Any, index: int) -> str:
    text = _safe_text(header_value)
    return text if text else f"column_{index}"


def _profile_sheet(sheet: Any) -> SheetProfile:
    row_count = int(sheet.max_row or 0)
    column_count = int(sheet.max_column or 0)
    header_values = next(
        sheet.iter_rows(min_row=1, max_row=1, values_only=True),
        (),
    )
    column_names = [
        _column_name(header_values[index - 1] if index - 1 < len(header_values) else None, index)
        for index in range(1, column_count + 1)
    ]

    numeric_counts = [0 for _ in range(column_count)]
    date_counts = [0 for _ in range(column_count)]
    non_empty_cell_count = 0

    for row in sheet.iter_rows(values_only=True):
        for index, value in enumerate(row[:column_count]):
            if value is None:
                continue
            if isinstance(value, str) and not value.strip():
                continue
            non_empty_cell_count += 1
            if _is_numeric(value):
                numeric_counts[index] += 1
            if _is_date_like(value):
                date_counts[index] += 1

    numeric_columns = [
        DetectedColumn(
            column_index=index + 1,
            column_letter=get_column_letter(index + 1),
            name=column_names[index],
            count=count,
        )
        for index, count in enumerate(numeric_counts)
        if count > 0
    ]
    date_like_columns = [
        DetectedColumn(
            column_index=index + 1,
            column_letter=get_column_letter(index + 1),
            name=column_names[index],
            count=count,
        )
        for index, count in enumerate(date_counts)
        if count > 0
    ]

    return SheetProfile(
        name=sheet.title,
        row_count=row_count,
        column_count=column_count,
        column_names=column_names,
        detected_numeric_columns=numeric_columns,
        detected_date_like_columns=date_like_columns,
        non_empty_cell_count=non_empty_cell_count,
    )


def _markdown_summary(profile: WorkbookProfile) -> str:
    lines = ["# Hospital Workbook Profile", ""]
    if not profile.found:
        lines.extend(["Workbook not found.", ""])
        for warning in profile.warnings:
            lines.append(f"- Warning: {warning}")
        return "\n".join(lines).rstrip() + "\n"

    lines.extend(
        [
            f"- Path: `{profile.path}`",
            f"- Source location: `{profile.source_location}`",
            f"- SHA256: `{profile.sha256}`",
            f"- Sheets: {len(profile.sheet_names)}",
            "",
            "| Sheet | Rows | Columns | Non-empty cells | Numeric columns | Date-like columns |",
            "| --- | ---: | ---: | ---: | ---: | ---: |",
        ]
    )
    for sheet in profile.sheets:
        lines.append(
            f"| {sheet.name} | {sheet.row_count} | {sheet.column_count} | "
            f"{sheet.non_empty_cell_count} | {len(sheet.detected_numeric_columns)} | "
            f"{len(sheet.detected_date_like_columns)} |"
        )
    if profile.warnings:
        lines.extend(["", "## Warnings"])
        lines.extend(f"- {warning}" for warning in profile.warnings)
    return "\n".join(lines).rstrip() + "\n"


def inspect_hospital_workbook(
    repo_root: str | Path = ".",
    output_json: str | Path = DEFAULT_WORKBOOK_PROFILE_JSON,
    output_markdown: str | Path = DEFAULT_WORKBOOK_PROFILE_MD,
) -> WorkbookProfile:
    """Inspect the hospital workbook structure and write ignored profile files."""

    root = Path(repo_root).resolve()
    workbook_path, source_location, profile_warnings = locate_hospital_workbook(root)

    if workbook_path is None:
        profile = WorkbookProfile(found=False, warnings=profile_warnings)
    else:
        with warning_lib.catch_warnings(record=True) as caught_warnings:
            warning_lib.simplefilter("always", UserWarning)
            workbook = load_workbook(workbook_path, read_only=True, data_only=False)
            sheet_names = list(workbook.sheetnames)
            sheets = [_profile_sheet(workbook[sheet_name]) for sheet_name in sheet_names]
            workbook.close()
        profile_warnings.extend(sorted({str(warning.message) for warning in caught_warnings}))
        profile = WorkbookProfile(
            found=True,
            path=str(workbook_path.relative_to(root)),
            source_location=source_location,
            sha256=sha256_file(workbook_path),
            sheet_names=sheet_names,
            sheets=sheets,
            warnings=profile_warnings,
        )

    json_path = root / output_json
    markdown_path = root / output_markdown
    json_path.parent.mkdir(parents=True, exist_ok=True)
    markdown_path.parent.mkdir(parents=True, exist_ok=True)
    profile.output_json = str(json_path.relative_to(root))
    profile.output_markdown = str(markdown_path.relative_to(root))
    json_path.write_text(
        json.dumps(profile.model_dump(mode="json"), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    markdown_path.write_text(_markdown_summary(profile), encoding="utf-8")
    return profile
