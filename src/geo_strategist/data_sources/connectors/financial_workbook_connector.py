"""Connector for the real hospital financial/cost-assumption workbook.

Loads `.data/manual/hospital_cf_workbook/hospital_rough_cf_payback_model_tokyo_aichi_osaka_beds_updated.xlsx`
and joins `hospital_master_68`, `cf_payback_model_68`, and
`bed_counts_official_68` by `master_id`. Evidence grades are derived
deterministically from columns already present in the workbook
(`data_basis`, `bed_verification_status`) — never guessed.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl


DEFAULT_PATH = Path(".data/manual/hospital_cf_workbook/hospital_rough_cf_payback_model_tokyo_aichi_osaka_beds_updated.xlsx")

SHEET_HOSPITAL_MASTER = "hospital_master_68"
SHEET_CF_PAYBACK_MODEL = "cf_payback_model_68"
SHEET_BED_COUNTS = "bed_counts_official_68"
SHEET_SOURCE_CATALOG = "source_catalog"


def _read_sheet_rows(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        return []
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue
        rows.append(dict(zip(header, row)))
    return rows


def _financial_evidence_grade(row: dict[str, Any]) -> str:
    if row.get("data_basis") == "施設別実績" and row.get("actual_revenue_JPY_mm") is not None:
        return "verified_source"
    return "model_estimate"


def _beds_evidence_grade(row: dict[str, Any]) -> str:
    status = row.get("verification_status")
    if status == "確認済":
        return "verified_source"
    if status:
        return "third_party_estimate"
    return "unverified_candidate"


def load_records(repo_root: str | Path = ".", *, path: str | Path | None = None) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Load one joined record per master_id. Returns (records, issues)."""

    repo_root = Path(repo_root).resolve()
    resolved = Path(path) if path else repo_root / DEFAULT_PATH
    if not resolved.is_absolute():
        resolved = repo_root / resolved
    if not resolved.exists():
        return [], [{
            "issue_code": "financial_workbook_source_missing",
            "severity": "error",
            "message": f"Expected source file not found: {resolved}",
        }]

    try:
        workbook = openpyxl.load_workbook(resolved, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - defensive against corrupt workbook
        return [], [{
            "issue_code": "financial_workbook_unreadable",
            "severity": "error",
            "message": str(exc),
        }]

    master_rows = {row["master_id"]: row for row in _read_sheet_rows(workbook, SHEET_HOSPITAL_MASTER) if row.get("master_id")}
    cf_rows = {row["master_id"]: row for row in _read_sheet_rows(workbook, SHEET_CF_PAYBACK_MODEL) if row.get("master_id")}
    bed_rows = {row["master_id"]: row for row in _read_sheet_rows(workbook, SHEET_BED_COUNTS) if row.get("master_id")}

    issues: list[dict[str, Any]] = []
    records: list[dict[str, Any]] = []
    for master_id, master in master_rows.items():
        cf_row = cf_rows.get(master_id)
        bed_row = bed_rows.get(master_id)
        if cf_row is None:
            issues.append({
                "issue_code": "financial_row_missing_for_master_id",
                "severity": "warning",
                "master_id": master_id,
            })
        record: dict[str, Any] = {
            "master_id": master_id,
            "prefecture": master.get("prefecture"),
            "hospital_name": master.get("hospital_name"),
            "operator_type": master.get("operator_type_normalized") or (cf_row or {}).get("operator_type"),
            "mhlw_source_url": master.get("mhlw_source_url"),
            "source_artifact": str(DEFAULT_PATH),
            "source_sheet": SHEET_HOSPITAL_MASTER,
            "source_record_id": master_id,
        }
        if cf_row:
            record.update({
                "data_basis": cf_row.get("data_basis"),
                "actual_revenue_JPY_mm": cf_row.get("actual_revenue_JPY_mm"),
                "actual_expenses_JPY_mm": cf_row.get("actual_expenses_JPY_mm"),
                "actual_ord_profit_JPY_mm": cf_row.get("actual_ord_profit_JPY_mm"),
                "beds_used_in_model": cf_row.get("beds_used_in_model"),
                "estimated_revenue_JPY_mm": cf_row.get("estimated_revenue_JPY_mm"),
                "expense_ratio": cf_row.get("expense_ratio"),
                "ebitda_cf_margin": cf_row.get("EBITDA_CF_margin"),
                "cash_expenses_JPY_mm": cf_row.get("cash_expenses_JPY_mm"),
                "hand_cf_JPY_mm": cf_row.get("hand_CF_JPY_mm"),
                "hand_cf_margin": cf_row.get("hand_CF_margin"),
                "construction_multiplier": cf_row.get("construction_multiplier"),
                "land_price_JPY_per_sqm": cf_row.get("land_price_JPY_per_sqm"),
                "construction_cost_JPY_mm": cf_row.get("construction_cost_JPY_mm"),
                "equipment_it_JPY_mm": cf_row.get("equipment_IT_JPY_mm"),
                "land_area_sqm": cf_row.get("land_area_sqm"),
                "land_cost_JPY_mm": cf_row.get("land_cost_JPY_mm"),
                "working_capital_JPY_mm": cf_row.get("working_capital_JPY_mm"),
                "contingency_JPY_mm": cf_row.get("contingency_JPY_mm"),
                "initial_investment_JPY_mm": cf_row.get("initial_investment_JPY_mm"),
                "payback_years": cf_row.get("payback_years"),
                "required_cf_for_target_payback_JPY_mm": cf_row.get("required_CF_for_target_payback_JPY_mm"),
                "cf_gap_JPY_mm": cf_row.get("CF_gap_JPY_mm"),
                "payback_flag": cf_row.get("payback_flag"),
                "primary_source_key": cf_row.get("primary_source_key"),
                "source_url": cf_row.get("source_url"),
                "model_note": cf_row.get("model_note"),
                "financial_evidence_grade": _financial_evidence_grade(cf_row),
                "land_construction_evidence_grade": "model_estimate",
            })
        if bed_row:
            record.update({
                "official_beds_total": bed_row.get("official_beds_total"),
                "general_beds": bed_row.get("general_beds"),
                "bed_source_type": bed_row.get("source_type"),
                "bed_source_url": bed_row.get("source_url"),
                "bed_verification_status": bed_row.get("verification_status"),
                "beds_evidence_grade": _beds_evidence_grade(bed_row),
            })
        records.append(record)
    return records, issues
