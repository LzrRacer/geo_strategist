"""S1 source registry for the site-selection pipeline.

Loads `configs/site_selection_sources.yaml`, resolves each declared source
against the local filesystem, and reports a `connected` / `not_configured` /
`missing_file` / `error` status per source. This module never performs a
live network call and never fabricates a record count or evidence grade: if
a declared source's file is missing, the entry fails closed with an
explicit issue instead of silently returning zero records.
"""

from __future__ import annotations

import json
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import yaml


DEFAULT_CONFIG_PATH = Path("configs/site_selection_sources.yaml")
OUTPUT_ROOT = Path(".runs/experiments/s1_source_registry")

VALID_STATUSES = ("connected", "not_configured", "missing_file", "error")


@dataclass(frozen=True)
class ResolvedSource:
    source_key: str
    category: str
    connector: str
    status: str
    evidence_grade: str | None
    path: str | None
    record_count: int | None
    description: str | None
    issue: str | None
    checked_at: str

    def to_dict(self) -> dict[str, Any]:
        return {
            "source_key": self.source_key,
            "category": self.category,
            "connector": self.connector,
            "status": self.status,
            "evidence_grade": self.evidence_grade,
            "path": self.path,
            "record_count": self.record_count,
            "description": self.description,
            "issue": self.issue,
            "checked_at": self.checked_at,
        }


@dataclass(frozen=True)
class SourceRegistryValidationResult:
    run_id: str
    output_dir: Path
    connected_count: int
    not_configured_count: int
    missing_or_error_count: int
    output_paths: dict[str, str]


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def load_source_registry_config(repo_root: str | Path = ".", config_path: str | Path = DEFAULT_CONFIG_PATH) -> dict[str, Any]:
    repo_root = Path(repo_root)
    cfg_path = Path(config_path)
    if not cfg_path.is_absolute():
        cfg_path = repo_root / cfg_path
    if not cfg_path.exists():
        return {"sources": []}
    with cfg_path.open("r", encoding="utf-8") as file:
        return yaml.safe_load(file) or {"sources": []}


def _count_jsonl_records(path: Path) -> int:
    if not path.exists():
        return 0
    count = 0
    with path.open("r", encoding="utf-8") as file:
        for line in file:
            if line.strip():
                count += 1
    return count


def _count_xlsx_sheet_records(path: Path, sheet_name: str) -> int | None:
    if not path.exists():
        return None
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:  # pragma: no cover - defensive against corrupt workbook
        return None
    if sheet_name not in workbook.sheetnames:
        return None
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    return sum(1 for row in rows if any(cell is not None for cell in row))


def resolve_source(repo_root: Path, entry: dict[str, Any]) -> ResolvedSource:
    checked_at = _now_iso()
    source_key = str(entry.get("source_key"))
    category = str(entry.get("category"))
    connector = str(entry.get("connector"))
    declared_status = entry.get("status", "not_configured")
    description = entry.get("description")
    evidence_grade = entry.get("default_evidence_grade")
    path_value = entry.get("path")

    if declared_status != "connected":
        return ResolvedSource(
            source_key=source_key,
            category=category,
            connector=connector,
            status="not_configured",
            evidence_grade=None,
            path=path_value,
            record_count=None,
            description=description,
            issue="source_not_configured_in_this_environment",
            checked_at=checked_at,
        )

    if entry.get("derived"):
        # Derived sources are computed deterministically from another
        # connected source at candidate/feature-generation time; there is
        # no separate file to resolve here.
        return ResolvedSource(
            source_key=source_key,
            category=category,
            connector=connector,
            status="connected",
            evidence_grade=evidence_grade,
            path=path_value,
            record_count=None,
            description=description,
            issue=None,
            checked_at=checked_at,
        )

    if not path_value:
        return ResolvedSource(
            source_key=source_key,
            category=category,
            connector=connector,
            status="error",
            evidence_grade=None,
            path=None,
            record_count=None,
            description=description,
            issue="connected_source_missing_path_in_config",
            checked_at=checked_at,
        )

    resolved_path = repo_root / path_value
    if not resolved_path.exists():
        return ResolvedSource(
            source_key=source_key,
            category=category,
            connector=connector,
            status="missing_file",
            evidence_grade=None,
            path=path_value,
            record_count=None,
            description=description,
            issue=f"configured source file not found: {path_value}",
            checked_at=checked_at,
        )

    fmt = entry.get("format")
    record_count: int | None
    if fmt == "jsonl":
        record_count = _count_jsonl_records(resolved_path)
    elif fmt == "xlsx":
        sheets = entry.get("sheets") or {}
        primary_sheet = sheets.get("cf_payback_model") or sheets.get("hospital_master")
        record_count = _count_xlsx_sheet_records(resolved_path, primary_sheet) if primary_sheet else None
    else:
        record_count = None

    return ResolvedSource(
        source_key=source_key,
        category=category,
        connector=connector,
        status="connected",
        evidence_grade=evidence_grade,
        path=path_value,
        record_count=record_count,
        description=description,
        issue=None,
        checked_at=checked_at,
    )


def resolve_all_sources(repo_root: str | Path = ".", config_path: str | Path = DEFAULT_CONFIG_PATH) -> list[ResolvedSource]:
    repo_root = Path(repo_root).resolve()
    config = load_source_registry_config(repo_root, config_path)
    return [resolve_source(repo_root, entry) for entry in config.get("sources", [])]


def get_source(source_key: str, repo_root: str | Path = ".", config_path: str | Path = DEFAULT_CONFIG_PATH) -> ResolvedSource | None:
    for source in resolve_all_sources(repo_root, config_path):
        if source.source_key == source_key:
            return source
    return None


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(json.dumps(row, ensure_ascii=False) for row in rows) + ("\n" if rows else ""), encoding="utf-8")


def _report_markdown(report: dict[str, Any]) -> str:
    lines = [
        "# S1 Source Registry Validation",
        "",
        f"Run ID: `{report['run_id']}`",
        f"Connected sources: {report['connected_count']}",
        f"Not-configured sources: {report['not_configured_count']}",
        f"Missing/error sources: {report['missing_or_error_count']}",
        "",
        "| source_key | category | status | evidence_grade | record_count |",
        "|---|---|---|---|---:|",
    ]
    for row in report["sources"]:
        lines.append(
            f"| `{row['source_key']}` | {row['category']} | `{row['status']}` | "
            f"{row['evidence_grade'] or '-'} | {row['record_count'] if row['record_count'] is not None else '-'} |"
        )
    lines.append("")
    return "\n".join(lines)


def run_source_registry_validation(
    repo_root: str | Path = ".",
    *,
    config_path: str | Path = DEFAULT_CONFIG_PATH,
    output_root: str | Path | None = None,
) -> SourceRegistryValidationResult:
    repo_root = Path(repo_root).resolve()
    run_id = str(uuid.uuid4())
    out_root = Path(output_root) if output_root else repo_root / OUTPUT_ROOT
    if not out_root.is_absolute():
        out_root = repo_root / out_root
    out_dir = out_root / run_id

    sources = resolve_all_sources(repo_root, config_path)
    rows = [source.to_dict() for source in sources]
    connected = sum(1 for row in rows if row["status"] == "connected")
    not_configured = sum(1 for row in rows if row["status"] == "not_configured")
    missing_or_error = sum(1 for row in rows if row["status"] in {"missing_file", "error"})

    output_paths = {
        "manifest": str(out_dir / "s1_manifest.json"),
        "source_registry_records": str(out_dir / "s1_source_registry_records.jsonl"),
        "report_json": str(out_dir / "s1_report.json"),
        "report_markdown": str(out_dir / "s1_report.md"),
    }
    report = {
        "run_id": run_id,
        "generated_at": _now_iso(),
        "connected_count": connected,
        "not_configured_count": not_configured,
        "missing_or_error_count": missing_or_error,
        "sources": rows,
    }
    manifest = {
        "run_id": run_id,
        "stage": "s1_source_registry",
        "config_path": str(config_path),
        "output_artifacts": {key: str(Path(path).relative_to(repo_root)) for key, path in output_paths.items()},
    }

    _write_json(Path(output_paths["manifest"]), manifest)
    _write_jsonl(Path(output_paths["source_registry_records"]), rows)
    _write_json(Path(output_paths["report_json"]), report)
    Path(output_paths["report_markdown"]).write_text(_report_markdown(report), encoding="utf-8")

    return SourceRegistryValidationResult(
        run_id=run_id,
        output_dir=out_dir,
        connected_count=connected,
        not_configured_count=not_configured,
        missing_or_error_count=missing_or_error,
        output_paths=output_paths,
    )
